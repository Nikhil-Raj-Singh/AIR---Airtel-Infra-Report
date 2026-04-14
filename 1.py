# ==========================================================
# Trend Dashboard v2.0 - Site Run Hours Analysis
# Following AIR v2.0 Architecture (Auth, Storage, Styling)
# OPTIMIZED WITH POLARS & STREAMLIT CACHING
# ==========================================================

import os
import sys
import subprocess
import re
import json
import pandas as pd
import numpy as np
import streamlit as st
import plotly.graph_objects as go
import plotly.express as px
from datetime import datetime
from io import BytesIO
import pyarrow.parquet as pq
import pyarrow as pa
import polars as pl  # Added Polars for high-speed aggregations
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Import custom modules
from config import COLORS, ADMIN_CREDENTIALS, ROLES, PERMISSIONS, SESSION, get_timestamp
from styles import load_styles
from auth import init_auth_state, is_admin, is_authenticated, show_login_form, logout, get_username
from data_processor import DataProcessor, DataHandler
from components_trend import ColumnMapper, TrendMetricCard

# ==========================================================
# CACHED DATA LOADERS (Materialized Views)
# ==========================================================
@st.cache_data(show_spinner=False)
def get_cached_unified_master():
    """Loads the master dataset once and holds it in memory."""
    handler = DataHandler()
    return handler.load_unified_master()

@st.cache_data(show_spinner=False)
def get_cached_trend_data(year, month):
    """Loads specific month data once."""
    handler = DataHandler()
    return handler.load_trend_data(year, month)

# ==========================================================
# EXCEL EXPORT HELPER (Optimized with Polars)
# ==========================================================
@st.cache_data(show_spinner=False)
def create_beautiful_month_comparison_excel(df, use_months=None):
    """
    Creates beautiful 2-row header Excel using Polars for high-speed pivoting 
    and openpyxl's ws.append for instant data writing.
    """
    if df is None or len(df) == 0:
        return None

    wb = Workbook()
    ws = wb.active
    ws.title = "Trend Report"

    color_map = {
        'EB': 'FFE4002B',      # Red
        'DG': 'FF005A9C',      # Blue
        'BB': 'FFFFB703',      # Yellow
        'Solar': 'FF00A300',   # Green
        'Total': 'FF6C757D'    # Gray
    }

    # Convert Pandas to Polars for blazing fast operations
    pldf = pl.from_pandas(df)

    if use_months:
        pldf = pldf.filter(pl.col('Source_Month').is_in(use_months))

    unique_months = sorted(pldf['Source_Month'].unique().to_list())
    if not unique_months:
        return None

    base_cols = ['Site ID', 'Circle', 'MIB Name', 'Status']
    metrics = ['EB_Run_Avg', 'DG_Run_Avg', 'BB_Run_Avg', 'Solar_Run_Avg', 'Total_Run_Avg']

    # ===== HIGH-SPEED PIVOT WITH POLARS =====
    # 1. Melt the data
    melted = pldf.unpivot(
        index=base_cols + ['Source_Month'],
        on=[m for m in metrics if m in pldf.columns],
        variable_name='Metric',
        value_name='Value'
    )
    
    # 2. Create the target column names (e.g., EB_Run_Avg_Jan)
    melted = melted.with_columns(
        (pl.col('Metric') + "_" + pl.col('Source_Month')).alias('Pivot_Col')
    )
    
    # 3. Pivot back to wide format instantly
    pivoted = melted.pivot(
        values='Value',
        index=base_cols,
        columns='Pivot_Col',
        aggregate_function='first'
    )

    # ===== CREATE EXCEL HEADERS =====
    col_num = 1
    ws.row_dimensions[1].height = 25
    ws.row_dimensions[2].height = 20
    header_font = Font(bold=True, color="FFFFFF", size=11)

    # Base column headers
    for base_col in base_cols:
        cell = ws.cell(row=1, column=col_num)
        cell.value = base_col
        cell.font = header_font
        cell.fill = PatternFill(start_color="FF333333", end_color="FF333333", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        ws.cell(row=2, column=col_num).fill = PatternFill(start_color="FF333333", end_color="FF333333", fill_type="solid")
        col_num += 1

    ordered_cols = base_cols.copy()

    # Metric headers with merged cells
    for metric in metrics:
        span = len(unique_months)
        if span > 1:
            ws.merge_cells(start_row=1, start_column=col_num, end_row=1, end_column=col_num + span - 1)

        metric_cell = ws.cell(row=1, column=col_num)
        metric_cell.value = metric.replace('_', ' ')
        metric_cell.font = header_font

        metric_type = metric.split('_')[0]
        cell_color = color_map.get(metric_type, 'FF6C757D')
        metric_cell.fill = PatternFill(start_color=cell_color, end_color=cell_color, fill_type="solid")
        metric_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for i, month in enumerate(unique_months):
            month_cell = ws.cell(row=2, column=col_num + i)
            month_cell.value = month[:3]
            month_cell.font = Font(bold=True, color="FFFFFF", size=10)
            month_cell.fill = PatternFill(start_color=cell_color, end_color=cell_color, fill_type="solid")
            month_cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Track column order to align Polars output exactly with Excel headers
            ordered_cols.append(f"{metric}_{month}")

        col_num += span

    # ===== ALIGN DATA AND WRITE FAST =====
    # Ensure all required columns exist in the dataframe, fill missing with None
    for col in ordered_cols:
        if col not in pivoted.columns:
            pivoted = pivoted.with_columns(pl.lit(None).alias(col))
            
    # Reorder columns to match Excel perfectly
    pivoted = pivoted.select(ordered_cols)
    
    # Write rows instantly (bypasses cell-by-cell bottleneck)
    for row in pivoted.iter_rows():
        ws.append(row)

    # ===== Format columns =====
    ws.column_dimensions['A'].width = 13
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 10
    for col in range(5, col_num):
        ws.column_dimensions[get_column_letter(col)].width = 11

    ws.freeze_panes = 'E3'

    excel_bytes = BytesIO()
    wb.save(excel_bytes)
    excel_bytes.seek(0)
    return excel_bytes.getvalue()

# ==========================================================
# AUTO STREAMLIT LAUNCHER
# ===========================================================
def _inside_streamlit():
    try:
        from streamlit.runtime.scriptrunner import get_script_run_ctx
        return get_script_run_ctx() is not None
    except Exception:
        return False

if not _inside_streamlit():
    subprocess.run([sys.executable, "-m", "streamlit", "run", os.path.abspath(__file__)], check=False)
    sys.exit(0)

# ==========================================================
# PAGE CONFIGURATION & STATE
# ==========================================================
st.set_page_config(page_title="Trend Dashboard — Site Analysis", layout="wide", initial_sidebar_state="expanded")
load_styles()
init_auth_state()

if 'df_raw' not in st.session_state: st.session_state.df_raw = None
if 'df_processed' not in st.session_state: st.session_state.df_processed = None
if 'column_mapping' not in st.session_state: st.session_state.column_mapping = {}
if 'show_mapping_interface' not in st.session_state: st.session_state.show_mapping_interface = False
if 'upload_year' not in st.session_state: st.session_state.upload_year = datetime.now().year
if 'upload_month' not in st.session_state: st.session_state.upload_month = datetime.now().month
if 'current_page' not in st.session_state: st.session_state.current_page = 'Dashboard'

# ==========================================================
# SIDEBAR
# ==========================================================
with st.sidebar:
    st.markdown(f"""
    <div style='text-align: center; margin-bottom: 16px; background: linear-gradient(160deg, {COLORS.AIRTEL_RED} 0%, #8B0016 100%); padding: 22px 16px; border-radius: 14px; box-shadow: 0 4px 16px rgba(228, 0, 43, 0.2);'>
        <h1 style='margin: 0; color: #fff; font-size: 28px; font-weight: 900; letter-spacing: -0.5px; line-height: 1;'>TREND</h1>
        <p style='margin: 6px 0 0 0; color: rgba(255,255,255,0.7); font-size: 11px; font-weight: 500; letter-spacing: 1.5px; text-transform: uppercase;'>Site Analysis</p>
    </div>
    """, unsafe_allow_html=True)
    st.markdown("---")

    if is_authenticated():
        st.markdown(f"<h3 style='color: #FFB703; font-size: 13px; margin: 15px 0; font-weight: bold;'>👤 USER STATUS</h3>", unsafe_allow_html=True)
        username = get_username()
        col1, col2 = st.columns([3, 1])
        with col1: st.markdown(f"<span style='font-family: Arial, sans-serif; font-weight: 600;'>{username}</span>", unsafe_allow_html=True)
        with col2:
            if is_admin(): st.markdown(f"<span style='background: {COLORS.AIRTEL_RED}; color: white; padding: 2px 8px; border-radius: 3px; font-size: 11px; font-weight: bold;'>ADMIN</span>", unsafe_allow_html=True)
            else: st.markdown(f"<span style='background: #6c757d; color: white; padding: 2px 8px; border-radius: 3px; font-size: 11px; font-weight: bold;'>GUEST</span>", unsafe_allow_html=True)

        st.divider()
        st.markdown(f"<h3 style='color: #FFB703; font-size: 13px; margin: 15px 0; font-weight: bold;'>🧭 NAVIGATION</h3>", unsafe_allow_html=True)

        pages = ["Dashboard", "Trend Analysis"]
        if is_admin(): pages.extend(["Data Export", "Admin Panel"])

        for page in pages:
            icon = "📊" if page == "Dashboard" else "📈" if page == "Trend Analysis" else "📥" if page == "Data Export" else "⚙️"
            if st.button(f"{icon} {page}", key=f"nav_{page}", use_container_width=True):
                st.session_state.current_page = page
                st.rerun()

        st.divider()
        if st.button("🚪 Logout", use_container_width=True):
            logout()
            st.rerun()

# ==========================================================
# MAIN PAGE - LOGIN OR DASHBOARD
# ==========================================================
if not is_authenticated():
    st.title("📈 Trend Dashboard — Site Run Hours Analysis")
    st.markdown("Enterprise-grade energy run pattern analysis across 27k+ sites")
    st.markdown("---")
    st.markdown("### 🔐 Authentication Required")
    show_login_form()
    st.stop()

# ==========================================================
# PAGE ROUTING
# ==========================================================
if st.session_state.current_page == "Dashboard":
    st.title("📈 Trend Dashboard — Site Run Hours Analysis")
    st.markdown("Track and analyze energy run patterns across site infrastructure.")

    handler = DataHandler()
    existing_data = handler.list_trend_data()
    
    # Use cached loader
    df = get_cached_unified_master() if existing_data else None

    if df is not None and len(df) > 0:
        st.markdown("---")
        st.markdown("### 📂 Loaded Datasets (Chronological Order)")
        dataset_info = " → ".join([f"{item['month_name'][:3]} {item['year']}" for item in existing_data])
        st.info(f"📊 Datasets: {dataset_info} | Total: {len(existing_data)} month(s) | {len(df):,} total sites")

        st.markdown("---")
        st.markdown("### 📊 Month-on-Month Average Comparison")
        unique_months = df['Source_Month'].unique().tolist()

        if len(unique_months) > 1:
            st.markdown("#### 📈 Monthly Energy Source Averages")
            comparison_data = []
            for month in unique_months:
                month_data = df[df['Source_Month'] == month]
                comparison_data.append({
                    'Month': month,
                    'EB Avg (h)': pd.to_numeric(month_data['EB_Run_Avg'], errors='coerce').mean() if 'EB_Run_Avg' in month_data.columns else 0,
                    'DG Avg (h)': pd.to_numeric(month_data['DG_Run_Avg'], errors='coerce').mean() if 'DG_Run_Avg' in month_data.columns else 0,
                    'BB Avg (h)': pd.to_numeric(month_data['BB_Run_Avg'], errors='coerce').mean() if 'BB_Run_Avg' in month_data.columns else 0,
                    'Solar Avg (h)': pd.to_numeric(month_data['Solar_Run_Avg'], errors='coerce').mean() if 'Solar_Run_Avg' in month_data.columns else 0,
                })

            df_comparison = pd.DataFrame(comparison_data)
            col_viz1, col_viz2 = st.columns(2)

            with col_viz1:
                fig_trend = px.line(df_comparison, x='Month', y=['EB Avg (h)', 'DG Avg (h)', 'BB Avg (h)', 'Solar Avg (h)'], title="Energy Source Trends", markers=True)
                st.plotly_chart(fig_trend, use_container_width=True)
            with col_viz2:
                fig_bar = px.bar(df_comparison, x='Month', y=['EB Avg (h)', 'DG Avg (h)', 'BB Avg (h)', 'Solar Avg (h)'], title="Monthly Comparison", barmode='group')
                st.plotly_chart(fig_bar, use_container_width=True)
        else:
            col1, col2, col3, col4, col5 = st.columns(5)
            avg_eb = pd.to_numeric(df['EB_Run_Avg'], errors='coerce').mean() if 'EB_Run_Avg' in df.columns else 0
            avg_dg = pd.to_numeric(df['DG_Run_Avg'], errors='coerce').mean() if 'DG_Run_Avg' in df.columns else 0
            avg_solar = pd.to_numeric(df['Solar_Run_Avg'], errors='coerce').mean() if 'Solar_Run_Avg' in df.columns else 0
            with col3: TrendMetricCard.render("Avg EB", f"{avg_eb:.2f}h", "Per site", "default")
            with col4: TrendMetricCard.render("Avg DG", f"{avg_dg:.2f}h", "Per site", "default")
            with col5: TrendMetricCard.render("Avg Solar", f"{avg_solar:.2f}h", "Per site", "default")

        st.markdown("---")
        st.markdown("### 💾 Download Month-on-Month Comparison Report")
        
        # Uses cached, high-speed function
        excel_data = create_beautiful_month_comparison_excel(df)
        
        if excel_data:
            col_dl1, col_dl2 = st.columns(2)
            with col_dl1:
                st.download_button(
                    label="📥 Download Excel Report",
                    data=excel_data,
                    file_name=f"month_comparison_{datetime.now().strftime('%Y%m%d')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
            with col_dl2:
                st.info(f"📊 {len(existing_data)} months | {len(df)} sites")

    elif not is_admin(): st.info("⏳ No data available yet. Please ask an admin to upload data.")
    else: st.info("📤 Upload a data file above to get started.")

# ==========================================================
# ADMIN PANEL PAGE
# ==========================================================
elif st.session_state.current_page == "Admin Panel":
    st.title("⚙️ Admin Panel")
    if not is_admin(): st.error("❌ Access denied."); st.stop()

    st.markdown("---")
    st.markdown("### 📥 Upload New Data")
    uploaded_file = st.file_uploader("Select CSV or Excel file", type=["csv", "xlsx", "xls"], key="admin_uploader")

    if uploaded_file:
        processor = DataProcessor()
        try:
            df_raw = processor.load_data(uploaded_file)
            st.session_state.df_raw = df_raw
            st.session_state.show_mapping_interface = True
        except Exception as e:
            st.error(f"❌ Error loading file: {str(e)}")

    if st.session_state.show_mapping_interface and st.session_state.df_raw is not None:
        st.markdown("---")
        processor = DataProcessor()
        mapper = ColumnMapper(st.session_state.df_raw, processor)
        column_mapping, mapping_valid = mapper.show_mapping_interface()
        st.session_state.column_mapping = column_mapping

        st.markdown("### 📅 Set Time Period")
        col_year, col_month, _ = st.columns([2, 2, 2])
        with col_year: upload_year = st.number_input("Year", value=datetime.now().year)
        with col_month: upload_month = st.selectbox("Month", range(1, 13), index=datetime.now().month - 1)

        if st.button("✅ Process & Upload", type="primary"):
            if mapping_valid:
                with st.spinner("🔄 Processing and rebuilding caches..."):
                    handler = DataHandler()
                    df_processed = processor.filter_and_calculate(st.session_state.df_raw, column_mapping, upload_year, upload_month)
                    if len(df_processed) > 0:
                        handler.save_trend_data(df_processed, upload_year, upload_month)
                        
                        # CRITICAL: Clear the Streamlit cache so the dashboard knows to refresh!
                        st.cache_data.clear() 
                        
                        st.session_state.show_mapping_interface = False
                        st.success("✅ Data processed and database updated.")
                        st.rerun()

    st.markdown("---")
    handler = DataHandler()
    existing_data = handler.list_trend_data()
    if existing_data:
        for dataset in existing_data:
            col1, _, _, col4 = st.columns([3, 1, 1, 1])
            with col1: st.markdown(f"**{dataset['month_name']} {dataset['year']}**")
            with col4:
                if st.button("🗑️ Delete", key=f"del_{dataset['year']}_{dataset['month']}"):
                    handler.delete_trend_data(dataset['year'], dataset['month'])
                    st.cache_data.clear() # Clear cache on delete too
                    st.rerun()

# ==========================================================
# TREND ANALYSIS PAGE
# ==========================================================
elif st.session_state.current_page == "Trend Analysis":
    st.title("📈 Trend Analysis")
    
    # Use cached loader
    df_unified = get_cached_unified_master()

    if df_unified is not None and len(df_unified) > 0:
        st.markdown("---")
        col1, col2, col3 = st.columns(3)
        with col1: selected_circle = st.selectbox("Circle", ["All"] + sorted(df_unified['Circle'].fillna('Unknown').astype(str).unique().tolist()))
        with col2: 
            mibs = df_unified['MIB Name'] if selected_circle == "All" else df_unified[df_unified['Circle'] == selected_circle]['MIB Name']
            selected_mib = st.selectbox("MIB Name", ["All"] + sorted(mibs.fillna('Unknown').astype(str).unique().tolist()))
        with col3: selected_metric = st.selectbox("Metric", ['EB_Run_Avg', 'DG_Run_Avg', 'BB_Run_Avg', 'Solar_Run_Avg', 'Total_Run_Avg'])

        df_filtered = df_unified.copy()
        if selected_circle != "All": df_filtered = df_filtered[df_filtered['Circle'].astype(str) == selected_circle]
        if selected_mib != "All": df_filtered = df_filtered[df_filtered['MIB Name'].astype(str) == selected_mib]

        st.markdown("---")
        st.markdown("### 💾 Download Filtered Report")
        
        # Uses cached, high-speed function
        excel_data = create_beautiful_month_comparison_excel(df_filtered, use_months=df_filtered['Source_Month'].unique().tolist())
        
        if excel_data:
            st.download_button(
                label="📥 Download Filtered Excel",
                data=excel_data,
                file_name=f"filtered_{datetime.now().strftime('%Y%m%d')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

# ==========================================================
# DATA EXPORT PAGE
# ==========================================================
elif st.session_state.current_page == "Data Export":
    st.title("📥 Data Export")
    if not is_admin(): st.error("❌ Access denied."); st.stop()

    handler = DataHandler()
    existing_data = handler.list_trend_data()

    if existing_data:
        dataset_options = [f"{item['month_name']} {item['year']} ({item['rows']:,} sites)" for item in existing_data]
        selected_indices = st.multiselect("Select months to stack:", range(len(existing_data)), format_func=lambda i: dataset_options[i])

        if selected_indices:
            # Replaced slow Pandas stack with high-speed Polars stack
            pl_dfs = []
            for idx in selected_indices:
                data_info = existing_data[idx]
                df_temp = get_cached_trend_data(data_info['year'], data_info['month'])
                if df_temp is not None:
                    pldf = pl.from_pandas(df_temp).with_columns(
                        pl.lit(data_info['month_name']).alias('Source_Month'),
                        pl.lit(data_info['year']).alias('Source_Year')
                    )
                    pl_dfs.append(pldf)

            if pl_dfs:
                # Instant vertical stacking
                df_stacked = pl.concat(pl_dfs, how="vertical_relaxed")
                csv_bytes = df_stacked.write_csv()
                
                st.download_button(
                    label="📥 Download Stacked CSV",
                    data=csv_bytes,
                    file_name=f"stacked_sites_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv",
                    mime="text/csv"
                )



            if pl_dfs:
                # Instant vertical stacking via Polars
                df_stacked = pl.concat(pl_dfs, how="vertical_relaxed")
                csv_bytes = df_stacked.write_csv()
                
                col_dl1, col_dl2 = st.columns(2)
                
                with col_dl1:
                    st.download_button(
                        label="📥 Download Stacked CSV",
                        data=csv_bytes,
                        file_name=f"stacked_sites_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv",
                        mime="text/csv"
                    )
                    
                with col_dl2:
                    st.info(f"✅ {df_stacked.height:,} total rows | {len(selected_indices)} month(s)")

                st.markdown("---")
                st.markdown("### Preview of Stacked Data")
                # Convert the small 10-row preview back to pandas for guaranteed Streamlit rendering compatibility
                st.dataframe(df_stacked.head(10).to_pandas(), use_container_width=True)
    else:
        st.info("📭 No datasets available.")

# ==========================================================
# FOOTER
# ==========================================================
st.markdown("---")
st.markdown("""
<div style="text-align: center; color: #888; font-size: 0.85rem; margin-top: 30px;">
    <strong>Trend Dashboard v2.0</strong> | Powered by Streamlit & Airtel Data Intelligence<br>
    Following AIR Central BI v2.0 Architecture
</div>
""", unsafe_allow_html=True)

